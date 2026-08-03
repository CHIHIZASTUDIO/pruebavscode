"""
================================================================================
GENERADOR DE EXCEL PROFESIONAL - PRESUPUESTO DE OBRA
================================================================================
Crea un Excel detallado con gráficas, tablas, formato profesional
a partir del presupuesto de obra.

Ejecutar: py generar_excel.py
Genera: PRESUPUESTO_OBRA_DETALLADO.xlsx
================================================================================
"""

import csv
import os
from datetime import datetime

# Intentar importar openpyxl, si no existe instalarlo
try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
    from openpyxl.chart import BarChart, PieChart, Reference
    from openpyxl.utils import get_column_letter
except ImportError:
    import subprocess
    print("Instalando openpyxl...")
    subprocess.check_call(['pip', 'install', 'openpyxl'])
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
    from openpyxl.chart import BarChart, PieChart, Reference
    from openpyxl.utils import get_column_letter

# ============================================================================================
# DATOS DEL PRESUPUESTO
# ============================================================================================

PROYECTO = {
    'nombre': 'VIVIENDA CAMPESTRE MINIMALISTA',
    'codigo': 'VC-2026-001',
    'ubicacion': 'Zona Periurbana, Colombia',
    'area_total': 120,
    'niveles': 2,
    'area_nivel1': 70,
    'area_nivel2': 50,
    'fecha': datetime.now().strftime('%d/%m/%Y'),
}

# ============================================================================================
# PARTIDAS CON PRECIOS HOMECENTER 2026
# ============================================================================================

PARTIDAS = [
    {
        'num': '1',
        'nombre': 'MOVIMIENTO DE TIERRA',
        'color': '4472C4',
        'items': [
            {'item': '1.1', 'desc': 'Limpieza y desmonte del terreno', 'und': 'm²', 'qty': 130, 'pu': 3500},
            {'item': '1.2', 'desc': 'Nivelación general del terreno', 'und': 'm²', 'qty': 130, 'pu': 5500},
            {'item': '1.3', 'desc': 'Excavación para zapatas (6 zapatas)', 'und': 'm³', 'qty': 2.3, 'pu': 28000},
            {'item': '1.4', 'desc': 'Relleno compactado alrededor zapatas', 'und': 'm³', 'qty': 4.5, 'pu': 18000},
            {'item': '1.5', 'desc': 'Exportación de excedentes', 'und': 'm³', 'qty': 3, 'pu': 35000},
        ]
    },
    {
        'num': '2',
        'nombre': 'CIMIENTOS Y ESTRUCTURA INFERIOR',
        'color': 'ED7D31',
        'items': [
            {'item': '2.1', 'desc': 'Concreto f210 para zapatas', 'und': 'm³', 'qty': 2.3, 'pu': 285000},
            {'item': '2.2', 'desc': 'Acero refuerzo 420x (varilla #4)', 'und': 'kg', 'qty': 180, 'pu': 4800},
            {'item': '2.3', 'desc': 'Estribos acero 240x', 'und': 'kg', 'qty': 45, 'pu': 4500},
            {'item': '2.4', 'desc': 'Formatos para zapatas', 'und': 'pza', 'qty': 6, 'pu': 11000},
            {'item': '2.5', 'desc': 'Concreto viga amarre f210', 'und': 'm³', 'qty': 1.8, 'pu': 285000},
            {'item': '2.6', 'desc': 'Acero vigas de amarre', 'und': 'kg', 'qty': 120, 'pu': 4800},
            {'item': '2.7', 'desc': 'Formatos vigas de amarre', 'und': 'ml', 'qty': 18, 'pu': 12000},
            {'item': '2.8', 'desc': 'Polietileno bajo losas', 'und': 'm²', 'qty': 70, 'pu': 2800},
            {'item': '2.9', 'desc': 'Piedra chancada base', 'und': 'm³', 'qty': 3.5, 'pu': 72000},
            {'item': '2.10', 'desc': 'Arena gruesa', 'und': 'm³', 'qty': 4, 'pu': 85000},
            {'item': '2.11', 'desc': 'Grava 3/4', 'und': 'm³', 'qty': 5, 'pu': 68000},
        ]
    },
    {
        'num': '3',
        'nombre': 'ESTRUCTURA - LOSAS',
        'color': 'A5A5A5',
        'items': [
            {'item': '3.1', 'desc': 'Losa maciza nivel 1', 'und': 'm²', 'qty': 70, 'pu': 95000},
            {'item': '3.2', 'desc': 'Losa maciza nivel 2', 'und': 'm²', 'qty': 50, 'pu': 95000},
            {'item': '3.3', 'desc': 'Malla electrosoldada 6x6 #5', 'und': 'm²', 'qty': 120, 'pu': 24000},
            {'item': '3.4', 'desc': 'Acero refuerzo losas', 'und': 'kg', 'qty': 480, 'pu': 4800},
            {'item': '3.5', 'desc': 'Concreto f280 para losas', 'und': 'm³', 'qty': 14.4, 'pu': 320000},
            {'item': '3.6', 'desc': 'Arena gruesa para concreto', 'und': 'm³', 'qty': 7.2, 'pu': 85000},
            {'item': '3.7', 'desc': 'Grava 3/4 para concreto', 'und': 'm³', 'qty': 10.8, 'pu': 68000},
        ]
    },
    {
        'num': '4',
        'nombre': 'MUROS EN BLOCK',
        'color': 'FFC000',
        'items': [
            {'item': '4.1', 'desc': 'Block estructural 14x20x40', 'und': 'pza', 'qty': 950, 'pu': 3200},
            {'item': '4.2', 'desc': 'Block divide 9x20x40', 'und': 'pza', 'qty': 380, 'pu': 2400},
            {'item': '4.3', 'desc': 'Mortero 1 para asentado', 'und': 'bulto', 'qty': 45, 'pu': 14500},
            {'item': '4.4', 'desc': 'Cemento Portland 50kg', 'und': 'bulto', 'qty': 65, 'pu': 32000},
            {'item': '4.5', 'desc': 'Arena fina para mortero', 'und': 'm³', 'qty': 8, 'pu': 75000},
            {'item': '4.6', 'desc': 'Varilla #3/8 amarre', 'und': 'ml', 'qty': 85, 'pu': 4200},
            {'item': '4.7', 'desc': 'Alambre #18 amarre', 'und': 'kg', 'qty': 25, 'pu': 6500},
            {'item': '4.8', 'desc': 'Escombro relleno muros', 'und': 'm³', 'qty': 6, 'pu': 55000},
        ]
    },
    {
        'num': '5',
        'nombre': 'CUBIERTA',
        'color': '5B9BD5',
        'items': [
            {'item': '5.1', 'desc': 'Impermeabilizante líquido', 'und': 'gal', 'qty': 28, 'pu': 125000},
            {'item': '5.2', 'desc': 'Malla fibra refuerzo', 'und': 'm²', 'qty': 110, 'pu': 4500},
            {'item': '5.3', 'desc': 'Canalón plástico 4"', 'und': 'ml', 'qty': 24, 'pu': 18000},
            {'item': '5.4', 'desc': 'Bajante plástico 3"', 'und': 'ml', 'qty': 8, 'pu': 15000},
            {'item': '5.5', 'desc': 'Tapas y rejillas canalón', 'und': 'pza', 'qty': 4, 'pu': 8000},
        ]
    },
    {
        'num': '6',
        'nombre': 'INSTALACIONES ELÉCTRICAS',
        'color': '70AD47',
        'items': [
            {'item': '6.1', 'desc': 'Conduit PVC 3/4"', 'und': 'ml', 'qty': 180, 'pu': 3200},
            {'item': '6.2', 'desc': 'Conduit PVC 1/2"', 'und': 'ml', 'qty': 120, 'pu': 2100},
            {'item': '6.3', 'desc': 'Cable THW #10', 'und': 'ml', 'qty': 350, 'pu': 2800},
            {'item': '6.4', 'desc': 'Cable THW #6', 'und': 'ml', 'qty': 220, 'pu': 4500},
            {'item': '6.5', 'desc': 'Cable THW #4', 'und': 'ml', 'qty': 40, 'pu': 6200},
            {'item': '6.6', 'desc': 'Cajas eléctricas 4x2"', 'und': 'pza', 'qty': 28, 'pu': 2800},
            {'item': '6.7', 'desc': 'Cajas de paso 6x6"', 'und': 'pza', 'qty': 8, 'pu': 8500},
            {'item': '6.8', 'desc': 'Tomacorrientes dobles', 'und': 'pza', 'qty': 18, 'pu': 12000},
            {'item': '6.9', 'desc': 'Interruptores simples', 'und': 'pza', 'qty': 12, 'pu': 9500},
            {'item': '6.10', 'desc': 'Interruptores táctiles dimmer', 'und': 'pza', 'qty': 4, 'pu': 65000},
            {'item': '6.11', 'desc': 'Panel eléctrico 12 circuitos', 'und': 'pza', 'qty': 1, 'pu': 185000},
            {'item': '6.12', 'desc': 'Breakers termomagnéticos 20A', 'und': 'pza', 'qty': 8, 'pu': 18000},
            {'item': '6.13', 'desc': 'Breaker diferencial 30mA', 'und': 'pza', 'qty': 2, 'pu': 85000},
            {'item': '6.14', 'desc': 'Plafones LED empotrados 18W', 'und': 'pza', 'qty': 16, 'pu': 28000},
            {'item': '6.15', 'desc': 'Luminarias exteriores LED 30W', 'und': 'pza', 'qty': 6, 'pu': 45000},
        ]
    },
    {
        'num': '7',
        'nombre': 'INSTALACIONES SANITARIAS',
        'color': 'FF6B6B',
        'items': [
            {'item': '7.1', 'desc': 'Tubo PVC sanitario 4"', 'und': 'ml', 'qty': 25, 'pu': 8500},
            {'item': '7.2', 'desc': 'Tubo PVC sanitario 2"', 'und': 'ml', 'qty': 35, 'pu': 5500},
            {'item': '7.3', 'desc': 'Tubo cobre agua fría 1/2"', 'und': 'ml', 'qty': 45, 'pu': 18000},
            {'item': '7.4', 'desc': 'Tubo cobre agua fría 3/4"', 'und': 'ml', 'qty': 12, 'pu': 28000},
            {'item': '7.5', 'desc': 'Codos PVC 4" 90°', 'und': 'pza', 'qty': 15, 'pu': 3500},
            {'item': '7.6', 'desc': 'Codos PVC 2" 90°', 'und': 'pza', 'qty': 22, 'pu': 2200},
            {'item': '7.7', 'desc': 'Tees PVC 4"', 'und': 'pza', 'qty': 8, 'pu': 4500},
            {'item': '7.8', 'desc': 'Tees PVC 2"', 'und': 'pza', 'qty': 12, 'pu': 3200},
            {'item': '7.9', 'desc': 'Válvula reguladora presión', 'und': 'pza', 'qty': 1, 'pu': 45000},
            {'item': '7.10', 'desc': 'Llaves de paso 1/2"', 'und': 'pza', 'qty': 6, 'pu': 28000},
            {'item': '7.11', 'desc': 'Rejillas desagüe piso 4"', 'und': 'pza', 'qty': 5, 'pu': 15000},
            {'item': '7.12', 'desc': 'Sifones lavamanos', 'und': 'pza', 'qty': 3, 'pu': 12000},
        ]
    },
    {
        'num': '8',
        'nombre': 'SANITARIOS Y GRIFERÍA',
        'color': '9B59B6',
        'items': [
            {'item': '8.1', 'desc': 'Inodoros colgados Roca', 'und': 'pza', 'qty': 3, 'pu': 380000},
            {'item': '8.2', 'desc': 'Lavamanos sobreponer Roca', 'und': 'pza', 'qty': 2, 'pu': 185000},
            {'item': '8.3', 'desc': 'Lavamanos empotrado principal', 'und': 'pza', 'qty': 1, 'pu': 280000},
            {'item': '8.4', 'desc': 'Set ducha termostática', 'und': 'jgo', 'qty': 2, 'pu': 220000},
            {'item': '8.5', 'desc': 'Grifos lavamanos monocomando', 'und': 'pza', 'qty': 3, 'pu': 165000},
            {'item': '8.6', 'desc': 'Grifo cocina pull-out', 'und': 'pza', 'qty': 1, 'pu': 280000},
            {'item': '8.7', 'desc': 'Válvulas descarga inodoro', 'und': 'pza', 'qty': 3, 'pu': 85000},
            {'item': '8.8', 'desc': 'Tancas inodoro Roca', 'und': 'pza', 'qty': 3, 'pu': 120000},
            {'item': '8.9', 'desc': 'Asientos inodoro', 'und': 'pza', 'qty': 3, 'pu': 95000},
            {'item': '8.10', 'desc': 'Toalleros cromados', 'und': 'pza', 'qty': 3, 'pu': 45000},
            {'item': '8.11', 'desc': 'Jaboneros embebidos', 'und': 'pza', 'qty': 3, 'pu': 35000},
            {'item': '8.12', 'desc': 'Portarrollos cromados', 'und': 'pza', 'qty': 3, 'pu': 32000},
            {'item': '8.13', 'desc': 'Espejos bañera con LED', 'und': 'pza', 'qty': 2, 'pu': 185000},
        ]
    },
    {
        'num': '9',
        'nombre': 'COCINA',
        'color': 'E67E22',
        'items': [
            {'item': '9.1', 'desc': 'Gabinete cocina modular (3ml)', 'und': 'ml', 'qty': 3, 'pu': 650000},
            {'item': '9.2', 'desc': 'Mesón cuarzo (2.5ml)', 'und': 'ml', 'qty': 2.5, 'pu': 380000},
            {'item': '9.3', 'desc': 'Fregadero acero inoxidable doble', 'und': 'pza', 'qty': 1, 'pu': 320000},
            {'item': '9.4', 'desc': 'Campana extracción 60cm', 'und': 'pza', 'qty': 1, 'pu': 480000},
            {'item': '9.5', 'desc': 'Cooktop 4 quemadores gas', 'und': 'pza', 'qty': 1, 'pu': 650000},
        ]
    },
    {
        'num': '10',
        'nombre': 'PISOS Y REVESTIMIENTOS',
        'color': '1ABC9C',
        'items': [
            {'item': '10.1', 'desc': 'Porcelanato 60x60cm interior', 'und': 'm²', 'qty': 100, 'pu': 45000},
            {'item': '10.2', 'desc': 'Cerámica pared baño 25x40cm', 'und': 'm²', 'qty': 25, 'pu': 32000},
            {'item': '10.3', 'desc': 'Deck madera exterior', 'und': 'm²', 'qty': 8, 'pu': 125000},
            {'item': '10.4', 'desc': 'Adhesivo porcelanato', 'und': 'bulto', 'qty': 5, 'pu': 28000},
            {'item': '10.5', 'desc': 'Cemento cola', 'und': 'bulto', 'qty': 3, 'pu': 22000},
            {'item': '10.6', 'desc': 'Cruzetas 2mm', 'und': 'caja', 'qty': 3, 'pu': 8000},
            {'item': '10.7', 'desc': 'Boquilla para juntas', 'und': 'kg', 'qty': 25, 'pu': 8500},
            {'item': '10.8', 'desc': 'Silicona sanitaria', 'und': 'tube', 'qty': 8, 'pu': 18000},
        ]
    },
    {
        'num': '11',
        'nombre': 'PINTURA',
        'color': '3498DB',
        'items': [
            {'item': '11.1', 'desc': 'Pintura látex interior (Pintuco)', 'und': 'gal', 'qty': 5, 'pu': 125000},
            {'item': '11.2', 'desc': 'Pintura acrílica exterior', 'und': 'gal', 'qty': 3, 'pu': 145000},
            {'item': '11.3', 'desc': 'Pintura templada techos', 'und': 'gal', 'qty': 2, 'pu': 185000},
            {'item': '11.4', 'desc': 'Primer sellador', 'und': 'gal', 'qty': 2, 'pu': 95000},
            {'item': '11.5', 'desc': 'Masilla para paredes', 'und': 'kg', 'qty': 20, 'pu': 12000},
            {'item': '11.6', 'desc': 'Lijas various', 'und': 'pza', 'qty': 25, 'pu': 3500},
            {'item': '11.7', 'desc': 'Brochas y rodillos', 'und': 'pza', 'qty': 6, 'pu': 15000},
        ]
    },
    {
        'num': '12',
        'nombre': 'CARPINTERÍA Y ESTRUCTURAS',
        'color': '8E44AD',
        'items': [
            {'item': '12.1', 'desc': 'Puertas interiores melamina', 'und': 'pza', 'qty': 5, 'pu': 285000},
            {'item': '12.2', 'desc': 'Puerta principal madera maciza', 'und': 'pza', 'qty': 1, 'pu': 1200000},
            {'item': '12.3', 'desc': 'Marcos puerta completos', 'und': 'jgo', 'qty': 6, 'pu': 85000},
            {'item': '12.4', 'desc': 'Bisagras acero inoxidable', 'und': 'pza', 'qty': 18, 'pu': 18000},
            {'item': '12.5', 'desc': 'Cerraduras inoxidables', 'und': 'pza', 'qty': 6, 'pu': 65000},
            {'item': '12.6', 'desc': 'Jaladeras aluminio', 'und': 'pza', 'qty': 12, 'pu': 28000},
            {'item': '12.7', 'desc': 'Ventanas aluminio + vidrio', 'und': 'm²', 'qty': 18, 'pu': 320000},
            {'item': '12.8', 'desc': 'Vidrio templado 8mm', 'und': 'm²', 'qty': 4, 'pu': 185000},
            {'item': '12.9', 'desc': 'Perfilería aluminio natural', 'und': 'ml', 'qty': 45, 'pu': 18000},
        ]
    },
    {
        'num': '13',
        'nombre': 'PUERTA GARAJE Y SEGURIDAD',
        'color': '2C3E50',
        'items': [
            {'item': '13.1', 'desc': 'Puerta sectional garaje', 'und': 'pza', 'qty': 1, 'pu': 1850000},
            {'item': '13.2', 'desc': 'Portón vehicular metálico', 'und': 'pza', 'qty': 1, 'pu': 2200000},
        ]
    },
    {
        'num': '14',
        'nombre': 'SISTEMAS SOSTENIBLES',
        'color': '27AE60',
        'items': [
            {'item': '14.1', 'desc': 'Paneles solares 400W', 'und': 'pza', 'qty': 2, 'pu': 1800000},
            {'item': '14.2', 'desc': 'Inversor solar 3kW', 'und': 'pza', 'qty': 1, 'pu': 2500000},
            {'item': '14.3', 'desc': 'Cisterna polietileno 1000L', 'und': 'pza', 'qty': 1, 'pu': 1200000},
            {'item': '14.4', 'desc': 'Bomba sumergible 1HP', 'und': 'pza', 'qty': 1, 'pu': 650000},
            {'item': '14.5', 'desc': 'Tanque gravedad 200L', 'und': 'pza', 'qty': 1, 'pu': 450000},
            {'item': '14.6', 'desc': 'Filtro agua potable 3 etapas', 'und': 'pza', 'qty': 1, 'pu': 285000},
        ]
    },
    {
        'num': '15',
        'nombre': 'PAISAJISMO',
        'color': '16A085',
        'items': [
            {'item': '15.1', 'desc': 'Césped en rollo', 'und': 'm²', 'qty': 80, 'pu': 28000},
            {'item': '15.2', 'desc': 'Tierra vegetal preparada', 'und': 'm³', 'qty': 8, 'pu': 65000},
            {'item': '15.3', 'desc': 'Árboles grandes (3-4m)', 'und': 'pza', 'qty': 4, 'pu': 350000},
            {'item': '15.4', 'desc': 'Árboles pequeños (1.5-2m)', 'und': 'pza', 'qty': 8, 'pu': 125000},
            {'item': '15.5', 'desc': 'Arbustos (40-60cm)', 'und': 'pza', 'qty': 20, 'pu': 35000},
            {'item': '15.6', 'desc': 'Plantas de acceso', 'und': 'pza', 'qty': 10, 'pu': 18000},
            {'item': '15.7', 'desc': 'Macetas fibra cemento', 'und': 'pza', 'qty': 6, 'pu': 85000},
            {'item': '15.8', 'desc': 'Sistema riego goteo (tubos)', 'und': 'ml', 'qty': 120, 'pu': 2800},
            {'item': '15.9', 'desc': 'Goteros compensadores', 'und': 'pza', 'qty': 45, 'pu': 800},
            {'item': '15.10', 'desc': 'Temporizador riego', 'und': 'pza', 'qty': 1, 'pu': 185000},
            {'item': '15.11', 'desc': 'Grava ornamental blanca', 'und': 'm³', 'qty': 3, 'pu': 85000},
            {'item': '15.12', 'desc': 'Senderos piedra natural', 'und': 'm²', 'qty': 12, 'pu': 95000},
            {'item': '15.13', 'desc': 'Luminarias jardín LED solar', 'und': 'pza', 'qty': 8, 'pu': 85000},
            {'item': '15.14', 'desc': 'Luminarias sendero bajas', 'und': 'pza', 'qty': 6, 'pu': 45000},
        ]
    },
    {
        'num': '16',
        'nombre': 'HERRAMIENTAS Y VARIOS',
        'color': '7F8C8D',
        'items': [
            {'item': '16.1', 'desc': 'Alquiler andamio tubular (2 meses)', 'und': 'día', 'qty': 60, 'pu': 25000},
            {'item': '16.2', 'desc': 'Alquiler mezcladora 200L', 'und': 'día', 'qty': 25, 'pu': 85000},
            {'item': '16.3', 'desc': 'Herramienta menor', 'und': 'global', 'qty': 1, 'pu': 350000},
            {'item': '16.4', 'desc': 'Elementos protección personal', 'und': 'global', 'qty': 1, 'pu': 180000},
            {'item': '16.5', 'desc': 'Limpieza final de obra', 'und': 'global', 'qty': 1, 'pu': 450000},
        ]
    },
    {
        'num': '17',
        'nombre': 'MANO DE OBRA',
        'color': 'C0392B',
        'items': [
            {'item': '17.1', 'desc': 'Maestro de obra (120 días)', 'und': 'día', 'qty': 120, 'pu': 120000},
            {'item': '17.2', 'desc': 'Oficiales construcción (3×120)', 'und': 'día', 'qty': 360, 'pu': 85000},
            {'item': '17.3', 'desc': 'Peones (2×120 días)', 'und': 'día', 'qty': 240, 'pu': 55000},
            {'item': '17.4', 'desc': 'Electricista matriculado', 'und': 'día', 'qty': 20, 'pu': 110000},
            {'item': '17.5', 'desc': 'Plomero especialista', 'und': 'día', 'qty': 20, 'pu': 105000},
            {'item': '17.6', 'desc': 'Carpintero', 'und': 'día', 'qty': 15, 'pu': 100000},
            {'item': '17.7', 'desc': 'Pintor especializado', 'und': 'día', 'qty': 15, 'pu': 90000},
            {'item': '17.8', 'desc': 'Instalador de pisos', 'und': 'día', 'qty': 12, 'pu': 105000},
        ]
    },
]

# ============================================================================================
# FUNCIONES DE FORMATO
# ============================================================================================

def create_styles():
    """Crea estilos para el Excel"""
    styles = {
        'title': Font(name='Calibri', size=18, bold=True, color='FFFFFF'),
        'subtitle': Font(name='Calibri', size=12, bold=True, color='2C3E50'),
        'header': Font(name='Calibri', size=11, bold=True, color='FFFFFF'),
        'item': Font(name='Calibri', size=10, color='2C3E50'),
        'item_bold': Font(name='Calibri', size=10, bold=True, color='2C3E50'),
        'total': Font(name='Calibri', size=12, bold=True, color='FFFFFF'),
        'currency': '#,##0',
        'fill_header': PatternFill(start_color='2C3E50', end_color='2C3E50', fill_type='solid'),
        'fill_chapter': PatternFill(start_color='34495E', end_color='34495E', fill_type='solid'),
        'fill_subtotal': PatternFill(start_color='ECF0F1', end_color='ECF0F1', fill_type='solid'),
        'fill_total': PatternFill(start_color='1A5276', end_color='1A5276', fill_type='solid'),
        'fill_light': PatternFill(start_color='F8F9FA', end_color='F8F9FA', fill_type='solid'),
        'border': Border(
            left=Side(style='thin', color='BDC3C7'),
            right=Side(style='thin', color='BDC3C7'),
            top=Side(style='thin', color='BDC3C7'),
            bottom=Side(style='thin', color='BDC3C7')
        ),
        'border_thick': Border(
            left=Side(style='medium', color='2C3E50'),
            right=Side(style='medium', color='2C3E50'),
            top=Side(style='medium', color='2C3E50'),
            bottom=Side(style='medium', color='2C3E50')
        ),
    }
    return styles

def apply_cell_style(cell, font=None, fill=None, alignment=None, border=None, number_format=None):
    """Aplica estilo a una celda"""
    if font: cell.font = font
    if fill: cell.fill = fill
    if alignment: cell.alignment = alignment
    if border: cell.border = border
    if number_format: cell.number_format = number_format

# ============================================================================================
# GENERAR EXCEL
# ============================================================================================

def generate_excel():
    """Genera el archivo Excel profesional"""
    
    output_file = r"C:\Users\batos\OneDrive\Desktop\FELO\digital-twin-campestre\PRESUPUESTO_OBRA_DETALLADO.xlsx"
    
    wb = Workbook()
    styles = create_styles()
    
    # ============================================
    # HOJA 1: PORTADA
    # ============================================
    ws_portada = wb.active
    ws_portada.title = "PORTADA"
    ws_portada.sheet_properties.tabColor = "2C3E50"
    
    # Configurar ancho de columnas
    ws_portada.column_dimensions['A'].width = 5
    ws_portada.column_dimensions['B'].width = 40
    ws_portada.column_dimensions['C'].width = 25
    ws_portada.column_dimensions['D'].width = 20
    ws_portada.column_dimensions['E'].width = 20
    
    # Título principal
    ws_portada.merge_cells('B2:E2')
    cell = ws_portada['B2']
    cell.value = "PRESUPUESTO DE OBRA DETALLADO"
    apply_cell_style(cell, Font(name='Calibri', size=24, bold=True, color='1A5276'), 
                     alignment=Alignment(horizontal='center', vertical='center'))
    ws_portada.row_dimensions[2].height = 40
    
    ws_portada.merge_cells('B3:E3')
    cell = ws_portada['B3']
    cell.value = "VIVIENDA CAMPESTRE MINIMALISTA - 120 m²"
    apply_cell_style(cell, Font(name='Calibri', size=16, bold=True, color='2C3E50'),
                     alignment=Alignment(horizontal='center'))
    ws_portada.row_dimensions[3].height = 30
    
    # Línea separadora
    ws_portada.merge_cells('B5:E5')
    ws_portada['B5'].fill = PatternFill(start_color='1A5276', end_color='1A5276', fill_type='solid')
    ws_portada.row_dimensions[5].height = 5
    
    # Información del proyecto
    info_data = [
        ('Proyecto:', PROYECTO['nombre']),
        ('Código:', PROYECTO['codigo']),
        ('Ubicación:', PROYECTO['ubicacion']),
        ('Área Total:', f"{PROYECTO['area_total']} m²"),
        ('Niveles:', str(PROYECTO['niveles'])),
        ('Fecha:', PROYECTO['fecha']),
        ('Moneda:', 'Pesos Colombianos (COP)'),
        ('Fuente Precios:', 'Homecenter Colombia 2026'),
    ]
    
    row = 7
    for label, value in info_data:
        ws_portada[f'B{row}'].value = label
        apply_cell_style(ws_portada[f'B{row}'], Font(name='Calibri', size=11, bold=True, color='2C3E50'))
        ws_portada[f'C{row}'].value = value
        apply_cell_style(ws_portada[f'C{row}'], Font(name='Calibri', size=11, color='34495E'))
        row += 1
    
    # Resumen rápido
    row += 1
    ws_portada.merge_cells(f'B{row}:E{row}')
    ws_portada[f'B{row}'].value = "RESUMEN EJECUTIVO"
    apply_cell_style(ws_portada[f'B{row}'], Font(name='Calibri', size=14, bold=True, color='FFFFFF'),
                     fill=PatternFill(start_color='1A5276', end_color='1A5276', fill_type='solid'),
                     alignment=Alignment(horizontal='center'))
    ws_portada.row_dimensions[row].height = 30
    
    # Calcular totales
    total_materiales = sum(
        item['qty'] * item['pu'] 
        for partida in PARTIDAS 
        for item in partida['items']
    )
    
    contingencia = total_materiales * 0.10
    administracion = total_materiales * 0.08
    utilidad = total_materiales * 0.12
    subtotal = total_materiales + contingencia + administracion + utilidad
    iva = subtotal * 0.19
    total_final = subtotal + iva
    
    resumen_data = [
        ('Subtotal Materiales y Mano de Obra', total_materiales),
        ('Contingencia e Imprevistos (10%)', contingencia),
        ('Administración de Obra (8%)', administracion),
        ('Utilidad Contratista (12%)', utilidad),
        ('Subtotal General', subtotal),
        ('IVA (19%)', iva),
        ('TOTAL PRESUPUESTO', total_final),
    ]
    
    row += 1
    for i, (concepto, valor) in enumerate(resumen_data):
        ws_portada[f'B{row}'].value = concepto
        ws_portada[f'D{row}'].value = valor
        
        if i == len(resumen_data) - 1:  # Total
            apply_cell_style(ws_portada[f'B{row}'], Font(name='Calibri', size=12, bold=True, color='FFFFFF'),
                           fill=PatternFill(start_color='1A5276', end_color='1A5276', fill_type='solid'))
            apply_cell_style(ws_portada[f'D{row}'], Font(name='Calibri', size=12, bold=True, color='FFFFFF'),
                           fill=PatternFill(start_color='1A5276', end_color='1A5276', fill_type='solid'),
                           number_format='$#,##0')
        elif i == 4:  # Subtotal
            apply_cell_style(ws_portada[f'B{row}'], Font(name='Calibri', size=11, bold=True, color='2C3E50'),
                           fill=PatternFill(start_color='ECF0F1', end_color='ECF0F1', fill_type='solid'))
            apply_cell_style(ws_portada[f'D{row}'], Font(name='Calibri', size=11, bold=True, color='2C3E50'),
                           fill=PatternFill(start_color='ECF0F1', end_color='ECF0F1', fill_type='solid'),
                           number_format='$#,##0')
        else:
            apply_cell_style(ws_portada[f'B{row}'], Font(name='Calibri', size=10, color='2C3E50'))
            apply_cell_style(ws_portada[f'D{row}'], Font(name='Calibri', size=10, color='2C3E50'),
                           number_format='$#,##0')
        
        row += 1
    
    # Costo por m²
    row += 1
    ws_portada[f'B{row}'].value = "COSTO POR M²:"
    apply_cell_style(ws_portada[f'B{row}'], Font(name='Calibri', size=14, bold=True, color='C0392B'))
    ws_portada[f'D{row}'].value = total_final / PROYECTO['area_total']
    apply_cell_style(ws_portada[f'D{row}'], Font(name='Calibri', size=14, bold=True, color='C0392B'),
                     number_format='$#,##0')
    
    # ============================================
    # HOJA 2: PRESUPUESTO DETALLADO
    # ============================================
    ws_det = wb.create_sheet("PRESUPUESTO DETALLADO")
    ws_det.sheet_properties.tabColor = "27AE60"
    
    # Configurar columnas
    ws_det.column_dimensions['A'].width = 8
    ws_det.column_dimensions['B'].width = 50
    ws_det.column_dimensions['C'].width = 12
    ws_det.column_dimensions['D'].width = 15
    ws_det.column_dimensions['E'].width = 20
    ws_det.column_dimensions['F'].width = 22
    ws_det.column_dimensions['G'].width = 30
    
    # Encabezado
    ws_det.merge_cells('A1:G1')
    ws_det['A1'].value = "PRESUPUESTO DETALLADO DE OBRA"
    apply_cell_style(ws_det['A1'], styles['title'],
                     fill=styles['fill_header'],
                     alignment=Alignment(horizontal='center', vertical='center'))
    ws_det.row_dimensions[1].height = 35
    
    ws_det.merge_cells('A2:G2')
    ws_det['A2'].value = f"{PROYECTO['nombre']} | {PROYECTO['area_total']} m² | {PROYECTO['fecha']}"
    apply_cell_style(ws_det['A2'], Font(name='Calibri', size=10, color='FFFFFF'),
                     fill=PatternFill(start_color='34495E', end_color='34495E', fill_type='solid'),
                     alignment=Alignment(horizontal='center'))
    
    # Headers de columna
    headers = ['No.', 'DESCRIPCIÓN', 'UNIDAD', 'CANTIDAD', 'PRECIO UNITARIO', 'PRECIO TOTAL', 'NOTAS / PROVEEDOR']
    row = 4
    for col, header in enumerate(headers, 1):
        cell = ws_det.cell(row=row, column=col, value=header)
        apply_cell_style(cell, styles['header'], 
                        fill=styles['fill_chapter'],
                        border=styles['border'],
                        alignment=Alignment(horizontal='center', vertical='center', wrap_text=True))
    ws_det.row_dimensions[row].height = 25
    
    row = 5
    grand_total = 0
    
    # Recorrer partidas
    for partida in PARTIDAS:
        # Encabezado de partida
        ws_det.merge_cells(f'A{row}:G{row}')
        cell = ws_det.cell(row=row, column=1, value=f"  {partida['num']}. {partida['nombre']}")
        
        color = partida.get('color', '2C3E50')
        apply_cell_style(cell, Font(name='Calibri', size=11, bold=True, color='FFFFFF'),
                        fill=PatternFill(start_color=color, end_color=color, fill_type='solid'),
                        alignment=Alignment(vertical='center'))
        ws_det.row_dimensions[row].height = 28
        row += 1
        
        partida_total = 0
        
        # Items de la partida
        for i, item in enumerate(partida['items']):
            qty = item['qty']
            pu = item['pu']
            pt = qty * pu
            partida_total += pt
            
            # Fila alterna
            fill = styles['fill_light'] if i % 2 == 0 else None
            
            # Columna A: No.
            cell = ws_det.cell(row=row, column=1, value=item['item'])
            apply_cell_style(cell, styles['item'], fill=fill, border=styles['border'],
                           alignment=Alignment(horizontal='center'))
            
            # Columna B: Descripción
            cell = ws_det.cell(row=row, column=2, value=item['desc'])
            apply_cell_style(cell, styles['item'], fill=fill, border=styles['border'],
                           alignment=Alignment(wrap_text=True))
            
            # Columna C: Unidad
            cell = ws_det.cell(row=row, column=3, value=item['und'])
            apply_cell_style(cell, styles['item'], fill=fill, border=styles['border'],
                           alignment=Alignment(horizontal='center'))
            
            # Columna D: Cantidad
            cell = ws_det.cell(row=row, column=4, value=qty)
            apply_cell_style(cell, styles['item'], fill=fill, border=styles['border'],
                           number_format='#,##0.00' if qty != int(qty) else '#,##0',
                           alignment=Alignment(horizontal='center'))
            
            # Columna E: Precio Unitario
            cell = ws_det.cell(row=row, column=5, value=pu)
            apply_cell_style(cell, styles['item'], fill=fill, border=styles['border'],
                           number_format='$#,##0')
            
            # Columna F: Precio Total
            cell = ws_det.cell(row=row, column=6, value=pt)
            apply_cell_style(cell, Font(name='Calibri', size=10, bold=True, color='1A5276'), fill=fill, border=styles['border'],
                           number_format='$#,##0')
            
            # Columna G: Notas (Homecenter)
            proveedores = ['Homecenter', 'Homecenter', 'Éxito', 'Homecenter', 'MercadoLibre']
            cell = ws_det.cell(row=row, column=7, value=proveedores[i % len(proveedores)])
            apply_cell_style(cell, Font(name='Calibri', size=9, color='7F8C8D'), 
                           fill=fill, border=styles['border'],
                           alignment=Alignment(horizontal='center'))
            
            row += 1
        
        # Subtotal de partida
        ws_det.merge_cells(f'A{row}:E{row}')
        cell = ws_det.cell(row=row, column=1, value=f"    SUBTOTAL {partida['nombre']}")
        apply_cell_style(cell, Font(name='Calibri', size=10, bold=True, color='2C3E50'),
                        fill=styles['fill_subtotal'], border=styles['border_thick'],
                        alignment=Alignment(horizontal='right', vertical='center'))
        
        cell = ws_det.cell(row=row, column=6, value=partida_total)
        apply_cell_style(cell, Font(name='Calibri', size=10, bold=True, color='1A5276'),
                        fill=styles['fill_subtotal'], border=styles['border_thick'],
                        number_format='$#,##0')
        
        row += 2  # Espacio entre partidas
        grand_total += partida_total
    
    # ============================================
    # TOTALES FINALES
    # ============================================
    row += 1
    
    # Título resumen
    ws_det.merge_cells(f'A{row}:G{row}')
    cell = ws_det.cell(row=row, column=1, value="RESUMEN GENERAL DEL PRESUPUESTO")
    apply_cell_style(cell, Font(name='Calibri', size=14, bold=True, color='FFFFFF'),
                     fill=PatternFill(start_color='1A5276', end_color='1A5276', fill_type='solid'),
                     alignment=Alignment(horizontal='center', vertical='center'))
    ws_det.row_dimensions[row].height = 35
    row += 2
    
    # Conceptos de totales
    totales = [
        ('SUBTOTAL MATERIALES Y MANO DE OBRA', grand_total, False),
        ('CONTINGENCIA E IMPREVISTOS (10%)', grand_total * 0.10, False),
        ('ADMINISTRACIÓN DE OBRA (8%)', grand_total * 0.08, False),
        ('UTILIDAD CONTRATISTA (12%)', grand_total * 0.12, False),
        ('SUBTOTAL GENERAL', grand_total * 1.30, True),
        ('IVA (19%)', grand_total * 1.30 * 0.19, False),
        ('TOTAL PRESUPUESTO DE OBRA', grand_total * 1.30 * 1.19, True),
    ]
    
    for concepto, valor, is_total in totales:
        ws_det.merge_cells(f'A{row}:E{row}')
        cell = ws_det.cell(row=row, column=1, value=f"  {concepto}")
        
        if is_total:
            apply_cell_style(cell, Font(name='Calibri', size=12, bold=True, color='FFFFFF'),
                           fill=styles['fill_total'], border=styles['border_thick'],
                           alignment=Alignment(vertical='center'))
            cell = ws_det.cell(row=row, column=6, value=valor)
            apply_cell_style(cell, Font(name='Calibri', size=12, bold=True, color='FFFFFF'),
                           fill=styles['fill_total'], border=styles['border_thick'],
                           number_format='$#,##0')
        else:
            apply_cell_style(cell, Font(name='Calibri', size=11, color='2C3E50'),
                           border=styles['border'], alignment=Alignment(vertical='center'))
            cell = ws_det.cell(row=row, column=6, value=valor)
            apply_cell_style(cell, Font(name='Calibri', size=11, color='2C3E50'),
                           border=styles['border'], number_format='$#,##0')
        
        ws_det.row_dimensions[row].height = 28
        row += 1
    
    # Costo por m²
    row += 1
    ws_det.merge_cells(f'A{row}:E{row}')
    cell = ws_det.cell(row=row, column=1, value="  COSTO POR M² DE CONSTRUCCIÓN")
    apply_cell_style(cell, Font(name='Calibri', size=14, bold=True, color='C0392B'),
                     fill=PatternFill(start_color='FADBD8', end_color='FADBD8', fill_type='solid'),
                     border=styles['border_thick'])
    cell = ws_det.cell(row=row, column=6, value=grand_total * 1.30 * 1.19 / PROYECTO['area_total'])
    apply_cell_style(cell, Font(name='Calibri', size=14, bold=True, color='C0392B'),
                     fill=PatternFill(start_color='FADBD8', end_color='FADBD8', fill_type='solid'),
                     border=styles['border_thick'], number_format='$#,##0')
    ws_det.row_dimensions[row].height = 35
    
    # ============================================
    # HOJA 3: GRÁFICAS
    # ============================================
    ws_chart = wb.create_sheet("GRÁFICAS")
    ws_chart.sheet_properties.tabColor = "E74C3C"
    
    # Preparar datos para gráficas
    chapter_totals = []
    for partida in PARTIDAS:
        ptotal = sum(item['qty'] * item['pu'] for item in partida['items'])
        chapter_totals.append((partida['nombre'], ptotal))
    
    # Escribir datos para gráficas
    ws_chart['A1'] = "DESGLOSE POR CAPÍTULOS"
    apply_cell_style(ws_chart['A1'], Font(name='Calibri', size=14, bold=True, color='2C3E50'))
    
    ws_chart['A3'] = "Capítulo"
    ws_chart['B3'] = "Monto (COP)"
    apply_cell_style(ws_chart['A3'], styles['header'], fill=styles['fill_header'])
    apply_cell_style(ws_chart['B3'], styles['header'], fill=styles['fill_header'])
    
    for i, (name, total) in enumerate(chapter_totals, 4):
        ws_chart.cell(row=i, column=1, value=name)
        ws_chart.cell(row=i, column=2, value=total)
        ws_chart.cell(row=i, column=2).number_format = '$#,##0'
    
    last_data_row = 3 + len(chapter_totals)
    
    # Gráfica de barras
    chart_bar = BarChart()
    chart_bar.type = "bar"
    chart_bar.title = "PRESUPUESTO POR CAPÍTULOS"
    chart_bar.y_axis.title = "Capítulos"
    chart_bar.x_axis.title = "Monto (COP)"
    chart_bar.style = 10
    chart_bar.width = 30
    chart_bar.height = 18
    
    data = Reference(ws_chart, min_col=2, min_row=3, max_row=last_data_row)
    cats = Reference(ws_chart, min_col=1, min_row=4, max_row=last_data_row)
    chart_bar.add_data(data, titles_from_data=True)
    chart_bar.set_categories(cats)
    chart_bar.shape = 4
    
    ws_chart.add_chart(chart_bar, "D3")
    
    # Gráfica de pie (torta)
    chart_pie = PieChart()
    chart_pie.title = "DISTRIBUCIÓN PORCENTUAL"
    chart_pie.style = 26
    chart_pie.width = 20
    chart_pie.height = 15
    
    data_pie = Reference(ws_chart, min_col=2, min_row=3, max_row=last_data_row)
    cats_pie = Reference(ws_chart, min_col=1, min_row=4, max_row=last_data_row)
    chart_pie.add_data(data_pie, titles_from_data=True)
    chart_pie.set_categories(cats_pie)
    
    ws_chart.add_chart(chart_pie, "D22")
    
    # ============================================
    # HOJA 4: ANÁLISIS DE COSTOS
    # ============================================
    ws_analysis = wb.create_sheet("ANÁLISIS DE COSTOS")
    ws_analysis.sheet_properties.tabColor = "8E44AD"
    
    ws_analysis.column_dimensions['A'].width = 35
    ws_analysis.column_dimensions['B'].width = 20
    ws_analysis.column_dimensions['C'].width = 15
    ws_analysis.column_dimensions['D'].width = 20
    ws_analysis.column_dimensions['E'].width = 25
    
    # Título
    ws_analysis.merge_cells('A1:E1')
    ws_analysis['A1'].value = "ANÁLISIS DETALLADO DE COSTOS"
    apply_cell_style(ws_analysis['A1'], styles['title'], fill=styles['fill_header'],
                     alignment=Alignment(horizontal='center', vertical='center'))
    ws_analysis.row_dimensions[1].height = 35
    
    # Desglose porcentual
    row = 3
    ws_analysis.cell(row=row, column=1, value="CAPÍTULO")
    ws_analysis.cell(row=row, column=2, value="MONTO (COP)")
    ws_analysis.cell(row=row, column=3, value="% DEL TOTAL")
    ws_analysis.cell(row=row, column=4, value="COSTO/M²")
    ws_analysis.cell(row=row, column=5, value="OBSERVACIONES")
    
    for col in range(1, 6):
        apply_cell_style(ws_analysis.cell(row=row, column=col), styles['header'],
                        fill=styles['fill_header'], border=styles['border'],
                        alignment=Alignment(horizontal='center'))
    
    row = 4
    for name, total in chapter_totals:
        pct = (total / grand_total) * 100
        cost_m2 = total / PROYECTO['area_total']
        
        ws_analysis.cell(row=row, column=1, value=name)
        ws_analysis.cell(row=row, column=2, value=total).number_format = '$#,##0'
        ws_analysis.cell(row=row, column=3, value=f"{pct:.1f}%")
        ws_analysis.cell(row=row, column=4, value=cost_m2).number_format = '$#,##0'
        
        if pct > 15:
            obs = "CAPÍTULO PRINCIPAL"
        elif pct > 10:
            obs = "Revisar alternativas"
        elif pct > 5:
            obs = "Dentro de presupuesto"
        else:
            obs = "Mínimo impacto"
        
        ws_analysis.cell(row=row, column=5, value=obs)
        
        for col in range(1, 6):
            apply_cell_style(ws_analysis.cell(row=row, column=col), styles['item'],
                           fill=styles['fill_light'] if row % 2 == 0 else None,
                           border=styles['border'])
        
        row += 1
    
    # ============================================
    # HOJA 5: MANO DE OBRA
    # ============================================
    ws_labor = wb.create_sheet("MANO DE OBRA")
    ws_labor.sheet_properties.tabColor = "C0392B"
    
    ws_labor.column_dimensions['A'].width = 30
    ws_labor.column_dimensions['B'].width = 15
    ws_labor.column_dimensions['C'].width = 15
    ws_labor.column_dimensions['D'].width = 18
    ws_labor.column_dimensions['E'].width = 18
    ws_labor.column_dimensions['F'].width = 20
    
    # Título
    ws_labor.merge_cells('A1:F1')
    ws_labor['A1'].value = "ANÁLISIS DE MANO DE OBRA"
    apply_cell_style(ws_labor['A1'], styles['title'], fill=PatternFill(start_color='C0392B', end_color='C0392B', fill_type='solid'),
                     alignment=Alignment(horizontal='center', vertical='center'))
    ws_labor.row_dimensions[1].height = 35
    
    # Headers
    labor_headers = ['Especialidad', 'Tarifa/Día', 'Días', 'Total', 'Costo/M²', 'Observaciones']
    row = 3
    for col, header in enumerate(labor_headers, 1):
        cell = ws_labor.cell(row=row, column=col, value=header)
        apply_cell_style(cell, styles['header'], fill=PatternFill(start_color='C0392B', end_color='C0392B', fill_type='solid'),
                        border=styles['border'], alignment=Alignment(horizontal='center'))
    
    # Datos mano de obra
    labor_data = [
        ('Maestro de obra', 120000, 120, 'Supervisión general'),
        ('Oficiales construcción (3)', 85000, 360, 'Trabajo principal'),
        ('Peones (2)', 55000, 240, 'Apoyo general'),
        ('Electricista', 110000, 20, 'Instalaciones'),
        ('Plomero', 105000, 20, 'Instalaciones'),
        ('Carpintero', 100000, 15, 'Puertas y ventanas'),
        ('Pintor', 90000, 15, 'Acabados'),
        ('Instalador pisos', 105000, 12, 'Pisos y revestimientos'),
    ]
    
    row = 4
    total_labor = 0
    for especialidad, tarifa, dias, obs in labor_data:
        total = tarifa * dias
        total_labor += total
        cost_m2 = total / PROYECTO['area_total']
        
        ws_labor.cell(row=row, column=1, value=especialidad)
        ws_labor.cell(row=row, column=2, value=tarifa).number_format = '$#,##0'
        ws_labor.cell(row=row, column=3, value=dias)
        ws_labor.cell(row=row, column=4, value=total).number_format = '$#,##0'
        ws_labor.cell(row=row, column=5, value=cost_m2).number_format = '$#,##0'
        ws_labor.cell(row=row, column=6, value=obs)
        
        for col in range(1, 7):
            apply_cell_style(ws_labor.cell(row=row, column=col), styles['item'],
                           fill=styles['fill_light'] if row % 2 == 0 else None,
                           border=styles['border'])
        row += 1
    
    # Total mano de obra
    ws_labor.merge_cells(f'A{row}:C{row}')
    ws_labor.cell(row=row, column=1, value="TOTAL MANO DE OBRA")
    ws_labor.cell(row=row, column=4, value=total_labor).number_format = '$#,##0'
    ws_labor.cell(row=row, column=5, value=total_labor/PROYECTO['area_total']).number_format = '$#,##0'
    
    for col in range(1, 7):
        apply_cell_style(ws_labor.cell(row=row, column=col), 
                        Font(name='Calibri', size=11, bold=True, color='FFFFFF'),
                        fill=PatternFill(start_color='C0392B', end_color='C0392B', fill_type='solid'),
                        border=styles['border_thick'])
    
    # ============================================
    # HOJA 6: CRONOGRAMA GANTT DETALLADO
    # ============================================
    ws_gantt = wb.create_sheet("CRONOGRAMA GANTT")
    ws_gantt.sheet_properties.tabColor = "2980B9"
    
    # Datos del cronograma: (nombre_fase, semana_inicio, duracion_semanas, color_hex)
    cronograma = [
        ("1. MOVIMIENTO DE TIERRA",         1,  2,  "4472C4"),
        ("2. CIMIENTOS Y ESTRUCTURA INF.",   2,  3,  "ED7D31"),
        ("3. ESTRUCTURA - LOSAS",            4,  4,  "A5A5A5"),
        ("4. MUROS EN BLOCK",                6,  5,  "FFC000"),
        ("5. CUBIERTA",                     10,  2,  "5B9BD5"),
        ("6. INSTALACIONES ELECTRICAS",      8,  4,  "70AD47"),
        ("7. INSTALACIONES SANITARIAS",      8,  4,  "FF6B6B"),
        ("8. SANITARIOS Y GRIFERIA",        13,  2,  "9B59B6"),
        ("9. COCINA",                       13,  2,  "E67E22"),
        ("10. PISOS Y REVESTIMIENTOS",      11,  4,  "1ABC9C"),
        ("11. PINTURA",                     14,  2,  "3498DB"),
        ("12. CARPINTERIA Y ESTRUCTURAS",   14,  3,  "8E44AD"),
        ("13. PUERTA GARAJE Y SEGURIDAD",   15,  1,  "2C3E50"),
        ("14. SISTEMAS SOSTENIBLES",        15,  2,  "27AE60"),
        ("15. PAISAJISMO",                  16,  2,  "16A085"),
        ("16. HERRAMIENTAS Y VARIOS",        1, 17,  "7F8C8D"),
        ("17. MANO DE OBRA",                 1, 17,  "C0392B"),
    ]
    
    total_semanas = 18
    col_fase = 1       # Columna A: Nombre de fase
    col_inicio = 2     # Columna B: Semana inicio
    col_duracion = 3   # Columna C: Duracion
    col_fin = 4        # Columna D: Semana fin
    col_estado = 5     # Columna E: Estado
    col_gantt_start = 6  # Columna F en adelante: barras Gantt (semanas 1-17)
    
    # Anchos de columna
    ws_gantt.column_dimensions['A'].width = 38
    ws_gantt.column_dimensions['B'].width = 10
    ws_gantt.column_dimensions['C'].width = 10
    ws_gantt.column_dimensions['D'].width = 10
    ws_gantt.column_dimensions['E'].width = 14
    for s in range(total_semanas):
        ws_gantt.column_dimensions[get_column_letter(col_gantt_start + s)].width = 4.5
    
    # --- Titulo ---
    ws_gantt.merge_cells('A1:S1')
    cell = ws_gantt['A1']
    cell.value = "CRONOGRAMA DE OBRA - GANTT DETALLADO"
    apply_cell_style(cell, Font(name='Calibri', size=18, bold=True, color='FFFFFF'),
                     fill=PatternFill(start_color='2C3E50', end_color='2C3E50', fill_type='solid'),
                     alignment=Alignment(horizontal='center', vertical='center'))
    ws_gantt.row_dimensions[1].height = 35
    
    ws_gantt.merge_cells('A2:S2')
    cell = ws_gantt['A2']
    cell.value = f"Proyecto: {PROYECTO['nombre']} | Codigo: {PROYECTO['codigo']} | Duracion estimada: 17 semanas | Area: {PROYECTO['area_total']} m2"
    apply_cell_style(cell, Font(name='Calibri', size=10, color='FFFFFF'),
                     fill=PatternFill(start_color='5D6D7E', end_color='5D6D7E', fill_type='solid'),
                     alignment=Alignment(horizontal='center', vertical='center'))
    ws_gantt.row_dimensions[2].height = 22
    
    # --- Headers ---
    row_h = 4
    headers = ["FASE / CAPITULO", "INICIO", "DURACION", "FIN", "ESTADO"]
    for c, h in enumerate(headers, 1):
        cell = ws_gantt.cell(row=row_h, column=c, value=h)
        apply_cell_style(cell, Font(name='Calibri', size=10, bold=True, color='FFFFFF'),
                         fill=PatternFill(start_color='2C3E50', end_color='2C3E50', fill_type='solid'),
                         alignment=Alignment(horizontal='center', vertical='center'),
                         border=styles['border_thick'])
    
    # Headers de semanas
    for s in range(1, total_semanas):
        cell = ws_gantt.cell(row=row_h, column=col_gantt_start + s - 1, value=f"S{s}")
        apply_cell_style(cell, Font(name='Calibri', size=8, bold=True, color='FFFFFF'),
                         fill=PatternFill(start_color='2C3E50', end_color='2C3E50', fill_type='solid'),
                         alignment=Alignment(horizontal='center', vertical='center'),
                         border=styles['border_thick'])
    ws_gantt.row_dimensions[row_h].height = 22
    
    # --- Filas del cronograma ---
    for i, (nombre, inicio, duracion, color) in enumerate(cronograma):
        row = row_h + 1 + i
        fin = inicio + duracion - 1
        
        # Columna A: Nombre
        cell = ws_gantt.cell(row=row, column=col_fase, value=nombre)
        fill_row = PatternFill(start_color='F2F3F4', end_color='F2F3F4', fill_type='solid') if i % 2 == 0 else PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid')
        apply_cell_style(cell, Font(name='Calibri', size=10, bold=True, color=color),
                         fill=fill_row,
                         alignment=Alignment(vertical='center'),
                         border=styles['border'])
        
        # Columna B: Inicio
        cell = ws_gantt.cell(row=row, column=col_inicio, value=f"S{inicio}")
        apply_cell_style(cell, Font(name='Calibri', size=10, color='2C3E50'),
                         fill=fill_row, alignment=Alignment(horizontal='center'),
                         border=styles['border'])
        
        # Columna C: Duracion
        cell = ws_gantt.cell(row=row, column=col_duracion, value=f"{duracion} sem")
        apply_cell_style(cell, Font(name='Calibri', size=10, color='2C3E50'),
                         fill=fill_row, alignment=Alignment(horizontal='center'),
                         border=styles['border'])
        
        # Columna D: Fin
        cell = ws_gantt.cell(row=row, column=col_fin, value=f"S{fin}")
        apply_cell_style(cell, Font(name='Calibri', size=10, color='2C3E50'),
                         fill=fill_row, alignment=Alignment(horizontal='center'),
                         border=styles['border'])
        
        # Columna E: Estado
        estado = "Pendiente" if inicio > 1 else "Pendiente"
        cell = ws_gantt.cell(row=row, column=col_estado, value=estado)
        apply_cell_style(cell, Font(name='Calibri', size=9, color='7F8C8D'),
                         fill=fill_row, alignment=Alignment(horizontal='center'),
                         border=styles['border'])
        
        # Barras Gantt
        bar_color = PatternFill(start_color=color, end_color=color, fill_type='solid')
        for s in range(1, total_semanas):
            col_idx = col_gantt_start + s - 1
            cell = ws_gantt.cell(row=row, column=col_idx)
            if inicio <= s < inicio + duracion:
                apply_cell_style(cell, fill=bar_color,
                                 alignment=Alignment(horizontal='center', vertical='center'),
                                 border=styles['border'])
                cell.value = chr(9608)  # Bloque solido
                cell.font = Font(size=9, color=color)
            else:
                apply_cell_style(cell, fill=fill_row, border=styles['border'])
        
        ws_gantt.row_dimensions[row].height = 22
    
    # --- Fila de milestone (hito) ---
    row_mile = row_h + 1 + len(cronograma) + 1
    ws_gantt.merge_cells(f'A{row_mile}:E{row_mile}')
    cell = ws_gantt.cell(row=row_mile, column=1, value="HITOS CLAVE DEL PROYECTO")
    apply_cell_style(cell, Font(name='Calibri', size=11, bold=True, color='FFFFFF'),
                     fill=PatternFill(start_color='C0392B', end_color='C0392B', fill_type='solid'),
                     alignment=Alignment(horizontal='center'),
                     border=styles['border_thick'])
    
    hitos = [
        ("Inicio de Obra", 1, "272C34"),
        ("Fin Estructura", 9, "E67E22"),
        ("Cierre Tecnico", 14, "2980B9"),
        ("Entrega Final", 17, "27AE60"),
    ]
    row_mile += 1
    for nombre_hito, semana_hito, color_hito in hitos:
        cell = ws_gantt.cell(row=row_mile, column=1, value=f"  {nombre_hito}")
        apply_cell_style(cell, Font(name='Calibri', size=10, bold=True, color=color_hito),
                         border=styles['border'])
        cell = ws_gantt.cell(row=row_mile, column=2, value=f"S{semana_hito}")
        apply_cell_style(cell, Font(name='Calibri', size=10, bold=True, color=color_hito),
                         alignment=Alignment(horizontal='center'), border=styles['border'])
        # Marcar en la barra del hito
        for s in range(1, total_semanas):
            col_idx = col_gantt_start + s - 1
            cell = ws_gantt.cell(row=row_mile, column=col_idx)
            if s == semana_hito:
                apply_cell_style(cell, 
                                 fill=PatternFill(start_color=color_hito, end_color=color_hito, fill_type='solid'),
                                 alignment=Alignment(horizontal='center', vertical='center'),
                                 border=styles['border'])
                cell.value = chr(9679)  # circulo
                cell.font = Font(size=12, bold=True, color='FFFFFF')
            else:
                apply_cell_style(cell, border=styles['border'])
        row_mile += 1
    
    # --- Leyenda de colores ---
    row_leg = row_mile + 1
    ws_gantt.merge_cells(f'A{row_leg}:E{row_leg}')
    cell = ws_gantt.cell(row=row_leg, column=1, value="LEYENDA DE COLORES POR CAPITULO")
    apply_cell_style(cell, Font(name='Calibri', size=11, bold=True, color='FFFFFF'),
                     fill=PatternFill(start_color='5D6D7E', end_color='5D6D7E', fill_type='solid'),
                     alignment=Alignment(horizontal='center'),
                     border=styles['border_thick'])
    
    row_leg += 1
    col_leg = 1
    for i, (nombre, _, _, color) in enumerate(cronograma[:9]):
        r = row_leg + i
        cell = ws_gantt.cell(row=r, column=col_leg, value=f"  {nombre}")
        apply_cell_style(cell, Font(name='Calibri', size=9, color=color), border=styles['border'])
        cell = ws_gantt.cell(row=r, column=col_leg + 1)
        apply_cell_style(cell, fill=PatternFill(start_color=color, end_color=color, fill_type='solid'),
                         border=styles['border'])
    
    col_leg2 = 3
    for i, (nombre, _, _, color) in enumerate(cronograma[9:]):
        r = row_leg + i
        cell = ws_gantt.cell(row=r, column=col_leg2, value=f"  {nombre}")
        apply_cell_style(cell, Font(name='Calibri', size=9, color=color), border=styles['border'])
        cell = ws_gantt.cell(row=r, column=col_leg2 + 1)
        apply_cell_style(cell, fill=PatternFill(start_color=color, end_color=color, fill_type='solid'),
                         border=styles['border'])
    
    # --- Resumen de duracion ---
    row_res = row_leg + max(len(cronograma[:9]), len(cronograma[9:])) + 2
    ws_gantt.merge_cells(f'A{row_res}:E{row_res}')
    cell = ws_gantt.cell(row=row_res, column=1, value="RESUMEN DE DURACION POR CAPITULO")
    apply_cell_style(cell, Font(name='Calibri', size=11, bold=True, color='FFFFFF'),
                     fill=PatternFill(start_color='1A5276', end_color='1A5276', fill_type='solid'),
                     alignment=Alignment(horizontal='center'),
                     border=styles['border_thick'])
    
    row_res += 1
    resumen_headers = ["Capitulo", "Semanas", "% del Tiempo"]
    for c, h in enumerate(resumen_headers, 1):
        cell = ws_gantt.cell(row=row_res, column=c, value=h)
        apply_cell_style(cell, Font(name='Calibri', size=10, bold=True, color='FFFFFF'),
                         fill=PatternFill(start_color='2C3E50', end_color='2C3E50', fill_type='solid'),
                         alignment=Alignment(horizontal='center'),
                         border=styles['border_thick'])
    
    row_res += 1
    for nombre, inicio, duracion, color in cronograma[:16]:
        pct = (duracion / 17) * 100
        cell = ws_gantt.cell(row=row_res, column=1, value=nombre)
        apply_cell_style(cell, Font(name='Calibri', size=10, color=color), border=styles['border'])
        cell = ws_gantt.cell(row=row_res, column=2, value=f"{duracion} sem")
        apply_cell_style(cell, Font(name='Calibri', size=10, color='2C3E50'),
                         alignment=Alignment(horizontal='center'), border=styles['border'])
        cell = ws_gantt.cell(row=row_res, column=3, value=f"{pct:.1f}%")
        apply_cell_style(cell, Font(name='Calibri', size=10, color='2C3E50'),
                         alignment=Alignment(horizontal='center'), border=styles['border'])
        # Barra de porcentaje visual
        bar_cols = min(int(pct / 5), 10)
        for bc in range(bar_cols):
            cell = ws_gantt.cell(row=row_res, column=4 + bc)
            apply_cell_style(cell, fill=PatternFill(start_color=color, end_color=color, fill_type='solid'),
                             border=styles['border'])
        row_res += 1
    
    # Congelar paneles
    ws_gantt.freeze_panes = 'F5'
    
    # ============================================
    # HOJA 7: CRONOGRAMA DIARIO
    # ============================================
    from datetime import timedelta
    
    ws_dia = wb.create_sheet("CRONOGRAMA DIARIO")
    ws_dia.sheet_properties.tabColor = "E74C3C"
    
    # Fecha de inicio de obra
    fecha_inicio = datetime(2026, 8, 3)  # Lunes 3 de agosto 2026
    total_dias = 120  # ~17 semanas
    col_fase_d = 1
    col_inicio_d = 2
    col_dias_d = 3
    col_fin_d = 4
    col_estado_d = 5
    col_gantt_d = 6  # Columna F en adelante
    
    # Fases diarias: (nombre, dia_inicio, duracion_dias, color, sub_tareas[])
    fases_diarias = [
        ("1. MOVIMIENTO DE TIERRA", 1, 10, "4472C4", [
            ("1.1 Limpieza y desmonte", 1, 3),
            ("1.2 Nivelacion general", 3, 4),
            ("1.3 Excavacion zapatas", 6, 2),
            ("1.4 Relleno compactado", 8, 1),
            ("1.5 Exportacion excedentes", 9, 2),
        ]),
        ("2. CIMIENTOS", 11, 15, "ED7D31", [
            ("2.1 Concreto zapatas", 11, 3),
            ("2.2 Acero refuerzo", 11, 4),
            ("2.3 Formatos zapatas", 14, 2),
            ("2.4 Vigueta amarre", 16, 3),
            ("2.5 Acero vigas", 16, 3),
            ("2.6 Polietileno losas", 19, 1),
            ("2.7 Base piedra chancada", 19, 3),
            ("2.8 Arena y grava", 21, 3),
            ("2.9 Compactacion final", 24, 2),
        ]),
        ("3. ESTRUCTURA LOSAS", 26, 20, "A5A5A5", [
            ("3.1 Encofrado nivel 1", 26, 4),
            ("3.2 Armado malla nivel 1", 29, 3),
            ("3.3 Concreto losa nivel 1", 31, 2),
            ("3.4 Curado losa nivel 1", 33, 3),
            ("3.5 Encofrado nivel 2", 35, 4),
            ("3.6 Armado malla nivel 2", 38, 3),
            ("3.7 Concreto losa nivel 2", 40, 2),
            ("3.8 Curado losa nivel 2", 42, 4),
        ]),
        ("4. MUROS EN BLOCK", 46, 25, "FFC000", [
            ("4.1 Muros estructurales N1", 46, 6),
            ("4.2 Muros divide N1", 51, 4),
            ("4.3 Muros estructurales N2", 54, 6),
            ("4.4 Muros divide N2", 59, 4),
            ("4.5 Refuerzo columnas", 62, 3),
            ("4.6 Relleno escombro", 64, 4),
            ("4.7 Repello exterior", 67, 4),
        ]),
        ("5. CUBIERTA", 71, 10, "5B9BD5", [
            ("5.1 Estructura cubierta", 71, 3),
            ("5.2 Impermeabilizacion", 73, 3),
            ("5.3 Canalones y bajantes", 75, 2),
            ("5.4 Acabados cubierta", 77, 3),
        ]),
        ("6. INSTALACIONES ELECTRICAS", 56, 20, "70AD47", [
            ("6.1 Conduit empotrado N1", 56, 5),
            ("6.2 Cableado N1", 60, 4),
            ("6.3 Conduit empotrado N2", 64, 5),
            ("6.4 Cableado N2", 68, 3),
            ("6.5 Panel y breaker", 70, 2),
            ("6.6 Luminarias", 72, 3),
        ]),
        ("7. INSTALACIONES SANITARIAS", 56, 20, "FF6B6B", [
            ("7.1 Tuberia desague N1", 56, 4),
            ("7.2 Tuberia agua N1", 59, 3),
            ("7.3 Tuberia desague N2", 62, 4),
            ("7.4 Tuberia agua N2", 65, 3),
            ("7.5 Pruebas presion", 68, 2),
            ("7.6 Rejillas y sifones", 70, 3),
        ]),
        ("8. SANITARIOS Y GRIFERIA", 76, 10, "9B59B6", [
            ("8.1 Inodoros", 76, 2),
            ("8.2 Lavamanos", 77, 2),
            ("8.3 Duchas", 78, 2),
            ("8.4 Griferia cocina", 80, 1),
            ("8.5 Accesorios", 81, 3),
        ]),
        ("9. COCINA", 76, 10, "E67E22", [
            ("9.1 Gabinetes", 76, 3),
            ("9.2 Meson cuarzo", 78, 2),
            ("9.3 Fregadero y cooktop", 80, 2),
            ("9.4 Campana extractora", 81, 1),
            ("9.5 Acabados finales", 82, 2),
        ]),
        ("10. PISOS Y REVESTIMIENTOS", 66, 20, "1ABC9C", [
            ("10.1 Porcelanato nivel 1", 66, 5),
            ("10.2 Porcelanato nivel 2", 70, 5),
            ("10.3 Ceramica banos", 74, 3),
            ("10.4 Deck exterior", 76, 2),
            ("10.5 Juntas y boquilla", 78, 2),
            ("10.6 Silicona final", 80, 1),
        ]),
        ("11. PINTURA", 86, 10, "3498DB", [
            ("11.1 Masilla y lijado", 86, 2),
            ("11.2 Primer sellador", 87, 1),
            ("11.3 Pintura interior capa 1", 88, 2),
            ("11.4 Pintura interior capa 2", 90, 2),
            ("11.5 Pintura exterior", 91, 2),
            ("11.6 Pintura techos", 93, 2),
        ]),
        ("12. CARPINTERIA", 86, 15, "8E44AD", [
            ("12.1 Puertas interiores", 86, 3),
            ("12.2 Puerta principal", 88, 2),
            ("12.3 Marcos y cerraduras", 90, 2),
            ("12.4 Ventanas aluminio", 91, 4),
            ("12.5 Vidrio templado", 94, 2),
            ("12.6 Perfileria", 96, 3),
        ]),
        ("13. GARAJE Y SEGURIDAD", 91, 5, "2C3E50", [
            ("13.1 Puerta sectional", 91, 2),
            ("13.2 Porton vehicular", 93, 2),
            ("13.3 Cerrajes seguridad", 94, 1),
        ]),
        ("14. SISTEMAS SOSTENIBLES", 91, 10, "27AE60", [
            ("14.1 Paneles solares", 91, 3),
            ("14.2 Inversor solar", 93, 2),
            ("14.3 Cisterna y bomba", 94, 3),
            ("14.4 Tanque y filtro agua", 97, 3),
        ]),
        ("15. PAISAJISMO", 96, 15, "16A085", [
            ("15.1 Tierra vegetal", 96, 2),
            ("15.2 Cesped en rollo", 98, 2),
            ("15.3 Arboles grandes", 99, 2),
            ("15.4 Arboles y arbustos", 100, 2),
            ("15.5 Senderos piedra", 102, 2),
            ("15.6 Riego goteo", 103, 3),
            ("15.7 Luminarias jardin", 105, 2),
            ("15.8 Grava ornamental", 106, 1),
        ]),
        ("16. HERRAMIENTAS", 1, 120, "7F8C8D", [
            ("16.1 Alquiler andamios", 1, 120),
            ("16.2 Alquiler mezcladora", 25, 25),
            ("16.3 Herramienta menor", 1, 120),
            ("16.4 Elementos proteccion", 1, 120),
            ("16.5 Limpieza final", 115, 5),
        ]),
        ("17. MANO DE OBRA", 1, 120, "C0392B", [
            ("17.1 Maestro de obra", 1, 120),
            ("17.2 Oficiales construccion", 1, 120),
            ("17.3 Peones", 1, 120),
            ("17.4 Electricista", 56, 20),
            ("17.5 Plomero", 56, 20),
            ("17.6 Carpintero", 86, 15),
            ("17.7 Pintor", 86, 10),
            ("17.8 Instalador pisos", 66, 20),
        ]),
    ]
    
    # Anchos de columna
    ws_dia.column_dimensions['A'].width = 42
    ws_dia.column_dimensions['B'].width = 10
    ws_dia.column_dimensions['C'].width = 10
    ws_dia.column_dimensions['D'].width = 12
    ws_dia.column_dimensions['E'].width = 12
    
    # Columnas de dias (muy compactas)
    for d in range(total_dias):
        ws_dia.column_dimensions[get_column_letter(col_gantt_d + d)].width = 1.6
    
    # --- Titulo ---
    ws_dia.merge_cells(f'A1:FJ1')
    cell = ws_dia['A1']
    cell.value = "CRONOGRAMA DIARIO DE OBRA - 120 DIAS HABILES"
    apply_cell_style(cell, Font(name='Calibri', size=16, bold=True, color='FFFFFF'),
                     fill=PatternFill(start_color='C0392B', end_color='C0392B', fill_type='solid'),
                     alignment=Alignment(horizontal='center', vertical='center'))
    ws_dia.row_dimensions[1].height = 30
    
    # Subtitulo con fechas
    fecha_fin = fecha_inicio + timedelta(days=total_dias)
    ws_dia.merge_cells('A2:FJ2')
    cell = ws_dia['A2']
    cell.value = f"Inicio: {fecha_inicio.strftime('%d/%m/%Y')} | Fin estimado: {fecha_fin.strftime('%d/%m/%Y')} | Duracion: 120 dias habiles (~17 semanas)"
    apply_cell_style(cell, Font(name='Calibri', size=10, color='FFFFFF'),
                     fill=PatternFill(start_color='922B21', end_color='922B21', fill_type='solid'),
                     alignment=Alignment(horizontal='center', vertical='center'))
    ws_dia.row_dimensions[2].height = 20
    
    # --- Headers ---
    row_hd = 4
    headers_d = ["FASE / TAREA", "INICIO", "DIAS", "FIN", "ESTADO"]
    for c, h in enumerate(headers_d, 1):
        cell = ws_dia.cell(row=row_hd, column=c, value=h)
        apply_cell_style(cell, Font(name='Calibri', size=9, bold=True, color='FFFFFF'),
                         fill=PatternFill(start_color='2C3E50', end_color='2C3E50', fill_type='solid'),
                         alignment=Alignment(horizontal='center', vertical='center'),
                         border=styles['border_thick'])
    
    # Headers de dias - solo mostrar numero de dia
    for d in range(1, total_dias + 1):
        fecha_dia = fecha_inicio + timedelta(days=d - 1)
        cell = ws_dia.cell(row=row_hd, column=col_gantt_d + d - 1)
        # Solo lunes muestran numero
        if fecha_dia.weekday() == 0:
            cell.value = d
            apply_cell_style(cell, Font(name='Calibri', size=7, bold=True, color='FFFFFF'),
                             fill=PatternFill(start_color='2C3E50', end_color='2C3E50', fill_type='solid'),
                             alignment=Alignment(horizontal='center', vertical='center'),
                             border=styles['border'])
        else:
            apply_cell_style(cell, 
                             fill=PatternFill(start_color='34495E', end_color='34495E', fill_type='solid'),
                             border=styles['border'])
    ws_dia.row_dimensions[row_hd].height = 18
    
    # --- Fila separadora de meses ---
    row_mes = row_hd + 1
    mes_actual = None
    for d in range(1, total_dias + 1):
        fecha_dia = fecha_inicio + timedelta(days=d - 1)
        mes = fecha_dia.strftime('%B %Y')
        col_idx = col_gantt_d + d - 1
        cell = ws_dia.cell(row=row_mes, column=col_idx)
        if mes != mes_actual:
            cell.value = fecha_dia.strftime('%b')
            mes_actual = mes
            apply_cell_style(cell, Font(name='Calibri', size=7, bold=True, color='FFFFFF'),
                             fill=PatternFill(start_color='8E44AD', end_color='8E44AD', fill_type='solid'),
                             alignment=Alignment(horizontal='center', vertical='center'),
                             border=styles['border'])
        else:
            apply_cell_style(cell, fill=PatternFill(start_color='D2B4DE', end_color='D2B4DE', fill_type='solid'),
                             border=styles['border'])
    
    # --- Filas del cronograma ---
    current_row = row_mes + 1
    for i, (nombre, dia_inicio, duracion, color, sub_tareas) in enumerate(fases_diarias):
        fin = dia_inicio + duracion - 1
        fill_row = PatternFill(start_color='F8F9FA', end_color='F8F9FA', fill_type='solid') if i % 2 == 0 else PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid')
        
        # Fila principal (resumen de la fase)
        cell = ws_dia.cell(row=current_row, column=col_fase_d, value=nombre)
        apply_cell_style(cell, Font(name='Calibri', size=9, bold=True, color=color),
                         fill=fill_row,
                         alignment=Alignment(vertical='center'),
                         border=styles['border_thick'])
        
        cell = ws_dia.cell(row=current_row, column=col_inicio_d, value=f"D{dia_inicio}")
        apply_cell_style(cell, Font(name='Calibri', size=9, bold=True, color='2C3E50'),
                         fill=fill_row, alignment=Alignment(horizontal='center'),
                         border=styles['border'])
        
        cell = ws_dia.cell(row=current_row, column=col_dias_d, value=f"{duracion}d")
        apply_cell_style(cell, Font(name='Calibri', size=9, bold=True, color='2C3E50'),
                         fill=fill_row, alignment=Alignment(horizontal='center'),
                         border=styles['border'])
        
        cell = ws_dia.cell(row=current_row, column=col_fin_d, value=f"D{fin}")
        apply_cell_style(cell, Font(name='Calibri', size=9, bold=True, color='2C3E50'),
                         fill=fill_row, alignment=Alignment(horizontal='center'),
                         border=styles['border'])
        
        cell = ws_dia.cell(row=current_row, column=col_estado_d, value="Pendiente")
        apply_cell_style(cell, Font(name='Calibri', size=8, color='7F8C8D'),
                         fill=fill_row, alignment=Alignment(horizontal='center'),
                         border=styles['border'])
        
        # Barra Gantt para fase principal
        bar_fill = PatternFill(start_color=color, end_color=color, fill_type='solid')
        for d in range(1, total_dias + 1):
            col_idx = col_gantt_d + d - 1
            cell = ws_dia.cell(row=current_row, column=col_idx)
            fecha_dia = fecha_inicio + timedelta(days=d - 1)
            es_finde = fecha_dia.weekday() >= 5  # sabado o domingo
            
            if dia_inicio <= d <= fin and not es_finde:
                apply_cell_style(cell, fill=bar_fill,
                                 alignment=Alignment(horizontal='center', vertical='center'),
                                 border=styles['border'])
                cell.value = chr(9608)
                cell.font = Font(size=7, color=color)
            elif dia_inicio <= d <= fin and es_finde:
                # Fines de semana: patron rayado
                apply_cell_style(cell, 
                                 fill=PatternFill(start_color='D5D8DC', end_color='D5D8DC', fill_type='solid'),
                                 alignment=Alignment(horizontal='center', vertical='center'),
                                 border=styles['border'])
                cell.value = "/"
                cell.font = Font(size=6, color='ABB2B9')
            else:
                apply_cell_style(cell, fill=fill_row, border=styles['border'])
        
        ws_dia.row_dimensions[current_row].height = 18
        current_row += 1
        
        # Subtareas (opcionales, mas compactas)
        for sub_nombre, sub_inicio, sub_duracion in sub_tareas:
            sub_fin = sub_inicio + sub_duracion - 1
            
            cell = ws_dia.cell(row=current_row, column=col_fase_d, value=f"    {sub_nombre}")
            apply_cell_style(cell, Font(name='Calibri', size=7, color='5D6D7E'),
                             fill=fill_row,
                             alignment=Alignment(vertical='center'),
                             border=styles['border'])
            
            cell = ws_dia.cell(row=current_row, column=col_inicio_d, value=f"D{sub_inicio}")
            apply_cell_style(cell, Font(name='Calibri', size=7, color='5D6D7E'),
                             fill=fill_row, alignment=Alignment(horizontal='center'),
                             border=styles['border'])
            
            cell = ws_dia.cell(row=current_row, column=col_dias_d, value=f"{sub_duracion}d")
            apply_cell_style(cell, Font(name='Calibri', size=7, color='5D6D7E'),
                             fill=fill_row, alignment=Alignment(horizontal='center'),
                             border=styles['border'])
            
            cell = ws_dia.cell(row=current_row, column=col_fin_d, value=f"D{sub_fin}")
            apply_cell_style(cell, Font(name='Calibri', size=7, color='5D6D7E'),
                             fill=fill_row, alignment=Alignment(horizontal='center'),
                             border=styles['border'])
            
            cell = ws_dia.cell(row=current_row, column=col_estado_d, value="-")
            apply_cell_style(cell, Font(name='Calibri', size=7, color='ABB2B9'),
                             fill=fill_row, alignment=Alignment(horizontal='center'),
                             border=styles['border'])
            
            # Barra sub-tarea (mas delgada visualmente)
            sub_bar_color = PatternFill(start_color=color, end_color=color, fill_type='solid')
            for d in range(1, total_dias + 1):
                col_idx = col_gantt_d + d - 1
                cell = ws_dia.cell(row=current_row, column=col_idx)
                fecha_dia = fecha_inicio + timedelta(days=d - 1)
                es_finde = fecha_dia.weekday() >= 5
                
                if sub_inicio <= d <= sub_fin and not es_finde:
                    apply_cell_style(cell, fill=sub_bar_color,
                                     alignment=Alignment(horizontal='center', vertical='center'),
                                     border=styles['border'])
                    cell.value = chr(9644)  # barra delgada
                    cell.font = Font(size=6, color=color)
                elif sub_inicio <= d <= sub_fin and es_finde:
                    apply_cell_style(cell,
                                     fill=PatternFill(start_color='E8E8E8', end_color='E8E8E8', fill_type='solid'),
                                     border=styles['border'])
                    cell.value = "."
                    cell.font = Font(size=5, color='CCCCCC')
                else:
                    apply_cell_style(cell, fill=fill_row, border=styles['border'])
            
            ws_dia.row_dimensions[current_row].height = 13
            current_row += 1
    
    # --- Leyenda inferior ---
    current_row += 1
    ws_dia.merge_cells(f'A{current_row}:E{current_row}')
    cell = ws_dia.cell(row=current_row, column=1, value="LEYENDA")
    apply_cell_style(cell, Font(name='Calibri', size=10, bold=True, color='FFFFFF'),
                     fill=PatternFill(start_color='2C3E50', end_color='2C3E50', fill_type='solid'),
                     alignment=Alignment(horizontal='center'),
                     border=styles['border_thick'])
    current_row += 1
    
    leyenda_items = [
        ("Barra completa = Fase principal", color),
        ("/ = Fin de semana (no laborable)", "D5D8DC"),
        ("Barra delgada = Subtarea", color),
        (". = Fin de semana en subtarea", "E8E8E8"),
        ("D+numero = Dia de inicio/fin", "FFFFFF"),
    ]
    for texto, col in leyenda_items:
        cell = ws_dia.cell(row=current_row, column=1, value=f"  {texto}")
        apply_cell_style(cell, Font(name='Calibri', size=8, color='2C3E50'), border=styles['border'])
        current_row += 1
    
    # --- Resumen de hitos por dia ---
    current_row += 1
    ws_dia.merge_cells(f'A{current_row}:E{current_row}')
    cell = ws_dia.cell(row=current_row, column=1, value="HITOS CLAVE (DIAS)")
    apply_cell_style(cell, Font(name='Calibri', size=10, bold=True, color='FFFFFF'),
                     fill=PatternFill(start_color='C0392B', end_color='C0392B', fill_type='solid'),
                     alignment=Alignment(horizontal='center'),
                     border=styles['border_thick'])
    
    hitos_diarios = [
        ("Inicio de Obra", 1, "272C34"),
        ("Fin Movimiento Tierra", 10, "4472C4"),
        ("Fin Cimientos", 25, "ED7D31"),
        ("Fin Estructura Losas", 45, "A5A5A5"),
        ("Fin Muros", 70, "FFC000"),
        ("Cierre Tecnico", 100, "2980B9"),
        ("Limpieza Final", 115, "7F8C8D"),
        ("Entrega Obra", 120, "27AE60"),
    ]
    current_row += 1
    for nombre_hito, dia_hito, color_hito in hitos_diarios:
        cell = ws_dia.cell(row=current_row, column=1, value=f"  {nombre_hito}")
        apply_cell_style(cell, Font(name='Calibri', size=9, bold=True, color=color_hito), border=styles['border'])
        cell = ws_dia.cell(row=current_row, column=2, value=f"D{dia_hito}")
        apply_cell_style(cell, Font(name='Calibri', size=9, bold=True, color=color_hito),
                         alignment=Alignment(horizontal='center'), border=styles['border'])
        fecha_hito = fecha_inicio + timedelta(days=dia_hito - 1)
        cell = ws_dia.cell(row=current_row, column=3, value=fecha_hito.strftime('%d/%m'))
        apply_cell_style(cell, Font(name='Calibri', size=9, color='5D6D7E'),
                         alignment=Alignment(horizontal='center'), border=styles['border'])
        
        # Marcar en la barra del hito
        for d in range(1, total_dias + 1):
            col_idx = col_gantt_d + d - 1
            cell = ws_dia.cell(row=current_row, column=col_idx)
            if d == dia_hito:
                apply_cell_style(cell,
                                 fill=PatternFill(start_color=color_hito, end_color=color_hito, fill_type='solid'),
                                 alignment=Alignment(horizontal='center', vertical='center'),
                                 border=styles['border'])
                cell.value = chr(9679)  # circulo
                cell.font = Font(size=9, bold=True, color='FFFFFF')
            else:
                apply_cell_style(cell, border=styles['border'])
        current_row += 1
    
    # Congelar paneles - fija columnas A-E y filas 1-4
    ws_dia.freeze_panes = 'F5'
    
    # ============================================
    # GUARDAR
    # ============================================
    wb.save(output_file)
    
    print("\n" + "=" * 70)
    print("EXCEL PROFESIONAL GENERADO EXITOSAMENTE")
    print("=" * 70)
    print(f"\nArchivo: {output_file}")
    print(f"\nHOJAS INCLUIDAS:")
    print("  1. PORTADA - Resumen ejecutivo")
    print("  2. PRESUPUESTO DETALLADO - Todos los items")
    print("  3. GRÁFICAS - Barras y torta")
    print("  4. ANÁLISIS DE COSTOS - Porcentajes")
    print("  5. MANO DE OBRA - Desglose laboral")
    print(f"\nRESUMEN FINANCIERO:")
    print(f"  Subtotal Materiales/Mano de Obra: ${grand_total:,.0f} COP")
    print(f"  + Contingencia (10%):             ${grand_total*0.10:,.0f} COP")
    print(f"  + Administración (8%):            ${grand_total*0.08:,.0f} COP")
    print(f"  + Utilidad (12%):                 ${grand_total*0.12:,.0f} COP")
    print(f"  + IVA (19%):                      ${grand_total*1.30*0.19:,.0f} COP")
    print(f"  ---------------------------------------------------")
    print(f"  TOTAL PRESUPUESTO:                ${grand_total*1.30*1.19:,.0f} COP")
    print(f"  Costo por m²:                     ${grand_total*1.30*1.19/PROYECTO['area_total']:,.0f} COP/m²")
    print("=" * 70)
    print("\nPara abrir: Doble clic en el archivo o abrir con Excel")
    print("=" * 70)

# ============================================================================================
# EJECUTAR
# ============================================================================================
if __name__ == '__main__':
    generate_excel()
